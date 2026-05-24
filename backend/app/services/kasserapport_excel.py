"""Multi-terminal close → Mirabelle-format .xlsx.

Produces an Excel workbook that mirrors the canonical weekly closing
sheet Manoj photographed at Silberbauer/Mirabelle:

    | Daily till report                 |   Today          |
    | Date                              |   09.03.2026     |
    | Cash closing - till out           |       18.799,00 |
    | Money to bank                     |            0,00 |
    | Paid out - change/byttep          |            0,00 |
    | Paid in                           |            0,00 |
    | Cash opening - till in            |            0,00 |
    | Cash total                        |       18.799,00 |
    | Gift cards accepted (total)       |            0,00 |
    | Mobile Pay                        |            0,00 |
    | Dankort from term. 1              |       24.292,51 |
    | Teller from term. 1               |       31.455,04 |
    | Amex from term. 1                 |            0,00 |
    | Total term. 1                     |       55.747,55 |
    | ...                                                  |
    | Cards total                       |       92.111,65 |
    | Payments total                    |      110.910,65 |
    | Sales POS (incl tax)              |      100.292,54 |
    | Cash difference (+/-)             |      -10.618,11 |
    | Initials                          | Caro             |

The owner opens the file, sees the same structure they've used for
years, can email it to their investor / drop into their existing
weekly workbook, etc.

Why generate this layout from scratch instead of filling the
customer's actual template:
  • We don't yet have customer-confirmed templates (waiting on
    Manoj's old boss)
  • This single-day variant is a useful artifact regardless — owner
    can paste-into a weekly sheet, or just keep the daily files
  • Once a customer uploads their template, we'll add a separate
    "fill template" path that maps these same row labels to their
    cell positions

Multi-terminal aware: if `aggregated.terminals` has multiple entries,
each gets its own Dankort / Teller / Amex / Total block in row order.
Single-terminal closes get a single block.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Any
from app.utils.time import utc_now

logger = logging.getLogger("bonbox.kasserapport_excel")


def render_close_xlsx(
    *,
    aggregated: dict,
    business_name: str = "",
    date_label: str = "",
    currency: str = "DKK",
    business_profile: dict | None = None,
    bilagsnummer: str | None = None,
    user: Any = None,
) -> bytes:
    """Build the close as an .xlsx workbook. Returns raw bytes; never
    raises — falls back to a minimal error workbook on any internal
    failure so the caller's UX flow continues.

    L4 defense-in-depth (multi_terminal_close): when `user` is supplied,
    re-checks the entitlement before rendering. Router gate is primary;
    this is the multi-barrier backup. PermissionError on refusal — the
    router converts to 402. Unit tests omit `user`.
    """
    # L4 — defensive feature check. Same multi-barrier rationale as
    # render_close_pdf; if a future refactor strips the router gate the
    # service still refuses.
    if user is not None:
        from app.services.billing import has_feature
        if not has_feature(user, "multi_terminal_close"):
            raise PermissionError("multi_terminal_close required")

    try:
        return _render_close_xlsx(
            aggregated=aggregated or {},
            business_name=business_name or "",
            date_label=date_label or "",
            currency=currency or "DKK",
            business_profile=business_profile or {},
            bilagsnummer=bilagsnummer,
        )
    except Exception as e:  # noqa: BLE001
        logger.exception("render_close_xlsx failed: %s", e)
        return _render_error_xlsx(str(e))


def _render_close_xlsx(
    *,
    aggregated: dict,
    business_name: str,
    date_label: str,
    currency: str,
    business_profile: dict,
    bilagsnummer: str | None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    # Sheet name = business slug (max 31 chars, no special chars per Excel spec)
    sheet_name = "".join(c if c.isalnum() or c in " -_" else "_" for c in (business_name or "Kasserapport"))[:31] or "Kasserapport"
    ws.title = sheet_name

    # ─── Style palette (Copenhagen-clean, matching the PDF) ────────
    INK = "171717"
    MUTED = "6b7280"
    AMBER_BG = "fef3c7"
    AMBER_INK = "b45309"
    GREEN_BG = "f0fdf4"
    GREEN_INK = "15803d"
    DIVIDER = "e5e7eb"

    font_h_doc = Font(name="Calibri", size=9, color=MUTED)
    font_h_business = Font(name="Calibri", size=16, bold=True, color=INK)
    font_h_meta = Font(name="Calibri", size=10, color=INK)
    font_section = Font(name="Calibri", size=10, bold=True, color="15803d")
    font_label = Font(name="Calibri", size=10, color=INK)
    font_label_bold = Font(name="Calibri", size=10, bold=True, color=INK)
    font_value = Font(name="Calibri", size=10, color=INK)
    font_value_bold = Font(name="Calibri", size=10, bold=True, color=INK)
    font_footer = Font(name="Calibri", size=8, italic=True, color=MUTED)

    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    thin_bottom = Border(bottom=Side(style="hair", color=DIVIDER))

    # Column widths — left col holds long labels like
    # "Cash closing - till out" + indent for terminal rows.
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22

    row = 1

    # ─── Header ──────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="KASSERAPPORT").font = font_h_doc
    row += 1

    ws.cell(row=row, column=1, value=business_name or "BonBox").font = font_h_business
    ws.row_dimensions[row].height = 24
    row += 1

    cvr = (business_profile.get("org_number") or "").strip() if business_profile else ""
    addr_parts = []
    if business_profile:
        a = (business_profile.get("address") or "").strip()
        z = (business_profile.get("zipcode") or "").strip()
        c = (business_profile.get("city") or "").strip()
        if a:
            addr_parts.append(a)
        if z and c:
            addr_parts.append(f"{z} {c}")
        elif c:
            addr_parts.append(c)
    addr = ", ".join(addr_parts)
    meta_bits = []
    if cvr:
        meta_bits.append(f"CVR {cvr}")
    if addr:
        meta_bits.append(addr)
    if meta_bits:
        ws.cell(row=row, column=1, value="  ·  ".join(meta_bits)).font = Font(
            name="Calibri", size=9, color=MUTED
        )
        row += 1

    # Date / closer / bilag line
    closer = (aggregated.get("closed_by") or "—").strip()
    date_str, day_name = _danish_date_label(date_label)
    line_bits = [date_str]
    if day_name:
        line_bits.append(day_name)
    line_bits.append(f"Lukket af: {closer}")
    if bilagsnummer:
        line_bits.append(f"Bilag: {bilagsnummer}")
    ws.cell(row=row, column=1, value="  ·  ".join(line_bits)).font = font_h_meta
    row += 2  # blank separator

    # ─── KONTANT ────────────────────────────────────────────────
    row = _section_header(ws, row, "KONTANT", font_section)
    row = _row_pair(
        ws, row, "Cash closing - till out",
        aggregated.get("cash_closing"), currency,
        font_label, font_value, align_left, align_right, thin_bottom,
    )
    row = _row_pair(ws, row, "Money to bank",
                    aggregated.get("money_to_bank"), currency,
                    font_label, font_value, align_left, align_right, thin_bottom)
    row = _row_pair(ws, row, "Paid out - change/byttep",
                    aggregated.get("paid_out"), currency,
                    font_label, font_value, align_left, align_right, thin_bottom)
    row = _row_pair(ws, row, "Paid in",
                    aggregated.get("paid_in"), currency,
                    font_label, font_value, align_left, align_right, thin_bottom)
    row = _row_pair(ws, row, "Cash opening - till in",
                    aggregated.get("cash_opening"), currency,
                    font_label, font_value, align_left, align_right, thin_bottom)
    row = _row_pair(ws, row, "Cash total",
                    aggregated.get("cash_total"), currency,
                    font_label_bold, font_value_bold, align_left, align_right, None)
    row += 1

    # ─── ANDRE ──────────────────────────────────────────────────
    row = _section_header(ws, row, "ANDRE", font_section)
    row = _row_pair(ws, row, "Gift cards accepted (total)",
                    aggregated.get("gift_cards_total"), currency,
                    font_label, font_value, align_left, align_right, thin_bottom)
    row = _row_pair(ws, row, "Mobile Pay",
                    aggregated.get("mobilepay_total"), currency,
                    font_label, font_value, align_left, align_right, None)
    row += 1

    # ─── Per-terminal blocks ───────────────────────────────────
    terminals = aggregated.get("terminals") or []
    if terminals:
        row = _section_header(ws, row, "KORTBETALINGER PR. TERMINAL", font_section)
        for i, t in enumerate(terminals, start=1):
            t_name = t.get("terminal_name") or f"Terminal {i}"
            cell = ws.cell(row=row, column=1, value=f"{i}. {t_name}")
            cell.font = font_label_bold
            cell.alignment = align_left
            row += 1
            row = _row_pair(ws, row, f"Dankort from term. {i}",
                            t.get("dankort"), currency,
                            font_label, font_value, align_left, align_right, thin_bottom)
            row = _row_pair(ws, row, f"Teller from term. {i}",
                            t.get("teller"), currency,
                            font_label, font_value, align_left, align_right, thin_bottom)
            row = _row_pair(ws, row, f"Amex from term. {i}",
                            t.get("amex"), currency,
                            font_label, font_value, align_left, align_right, thin_bottom)
            row = _row_pair(ws, row, f"Total term. {i}",
                            t.get("total"), currency,
                            font_label_bold, font_value_bold, align_left, align_right, None)
            row += 1

    # ─── OPSUMMERING ────────────────────────────────────────────
    row = _section_header(ws, row, "OPSUMMERING", font_section)
    row = _row_pair(ws, row, "Cards total",
                    aggregated.get("cards_total"), currency,
                    font_label_bold, font_value_bold, align_left, align_right, thin_bottom)
    row = _row_pair(ws, row, "Payments total",
                    aggregated.get("payments_total"), currency,
                    font_label_bold, font_value_bold, align_left, align_right, None)
    row += 1

    # ─── MOMS ───────────────────────────────────────────────────
    payments_total = aggregated.get("payments_total")
    revenue = aggregated.get("revenue") or {}
    moms_subtotal = revenue.get("subtotal_excl_moms")
    moms_amount = revenue.get("moms_amount")
    moms_total = revenue.get("total_incl_moms")
    if moms_total is None and payments_total is not None:
        moms_total = payments_total
        moms_subtotal = payments_total / 1.25
        moms_amount = payments_total - moms_subtotal
        moms_estimate = True
    else:
        moms_estimate = False

    moms_label = "MOMS" + ("  (estimat fra payments_total ved 25%)" if moms_estimate else "")
    row = _section_header(ws, row, moms_label, font_section)
    row = _row_pair(ws, row, "Subtotal ekskl. moms",
                    moms_subtotal, currency,
                    font_label, font_value, align_left, align_right, thin_bottom)
    row = _row_pair(ws, row, "Moms 25%",
                    moms_amount, currency,
                    font_label, font_value, align_left, align_right, thin_bottom)
    row = _row_pair(ws, row, "Total inkl. moms",
                    moms_total, currency,
                    font_label_bold, font_value_bold, align_left, align_right, None)
    row += 1

    # ─── AFSTEMNING (with cash-diff highlight) ─────────────────
    row = _section_header(ws, row, "AFSTEMNING", font_section)
    row = _row_pair(ws, row, "Sales POS (incl. tax)",
                    aggregated.get("sales_pos"), currency,
                    font_label_bold, font_value_bold, align_left, align_right, thin_bottom)

    # Cash difference row — coloured fill
    cash_diff = aggregated.get("cash_difference") or 0
    diff_flagged = bool(aggregated.get("cash_diff_flagged"))
    fill_bg = AMBER_BG if diff_flagged else GREEN_BG
    fill_ink = AMBER_INK if diff_flagged else GREEN_INK

    label_cell = ws.cell(row=row, column=1, value="Cash difference (+/-)")
    label_cell.font = font_label_bold
    label_cell.alignment = align_left
    label_cell.fill = PatternFill(start_color=fill_bg, end_color=fill_bg, fill_type="solid")

    val_cell = ws.cell(row=row, column=2, value=_money_value(cash_diff))
    val_cell.font = Font(name="Calibri", size=10, bold=True, color=fill_ink)
    val_cell.alignment = align_right
    val_cell.fill = PatternFill(start_color=fill_bg, end_color=fill_bg, fill_type="solid")
    val_cell.number_format = '#,##0.00 "' + currency + '"'
    row += 1

    flagged_reason = aggregated.get("flagged_reason") or ""
    if diff_flagged and flagged_reason:
        cell = ws.cell(row=row, column=1, value="⚠ " + flagged_reason)
        cell.font = Font(name="Calibri", size=8, italic=True, color=AMBER_INK)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1
    row += 1

    # ─── Initials / signature row (matches Mirabelle's "Initials" row) ─
    row = _section_header(ws, row, "UNDERSKRIFT", font_section)
    ws.cell(row=row, column=1, value="Lukket af").font = font_label_bold
    ws.cell(row=row, column=2, value=closer).font = font_value
    row += 1
    ws.cell(row=row, column=1, value="Godkendt (ejer / leder)").font = font_label_bold
    ws.cell(row=row, column=2, value="").font = font_value
    row += 2

    # ─── Footer ─────────────────────────────────────────────────
    ws.cell(row=row, column=1, value=(
        f"Opbevares i 5 år iht. Bogføringsloven §10  ·  "
        f"Genereret af BonBox  ·  bonbox.dk  ·  "
        f"{utc_now().strftime('%d.%m.%Y %H:%M UTC')}"
    )).font = font_footer
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    # ─── Page setup so prints land cleanly ─────────────────────
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_options.horizontalCentered = False

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _section_header(ws, row: int, text: str, font) -> int:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = font
    return row + 1


def _row_pair(ws, row, label, value, currency, font_label, font_value,
              align_left, align_right, border) -> int:
    """Render a 2-col label/value row; numeric `value` gets number_format
    so Excel treats it as a real number (not a string), enabling SUM,
    pivot tables, and formula references in the customer's downstream
    workflow."""
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = font_label
    label_cell.alignment = align_left
    if border:
        label_cell.border = border

    val = _money_value(value)
    val_cell = ws.cell(row=row, column=2, value=val)
    val_cell.font = font_value
    val_cell.alignment = align_right
    if border:
        val_cell.border = border
    if val is not None:
        val_cell.number_format = '#,##0.00 "' + currency + '"'
    return row + 1


def _money_value(value):
    """Coerce input to float for openpyxl. Returns None for non-numeric
    so the cell stays empty (Excel renders blank, not "—")."""
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


# Day names in Danish (same logic as in kasserapport_pdf)
_DK_DAYS = ["Mandag", "Tirsdag", "Onsdag", "Torsdag", "Fredag", "Lørdag", "Søndag"]


def _danish_date_label(date_label: str) -> tuple[str, str]:
    if not date_label:
        d = utc_now()
        return (d.strftime("%d.%m.%Y"), _DK_DAYS[d.weekday()])
    if "(" in date_label and ")" in date_label:
        date_part = date_label.split("(")[0].strip()
        day_part = date_label.split("(")[1].split(")")[0].strip()
        return (date_part, day_part)
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y"):
        try:
            d = datetime.strptime(date_label.strip(), fmt)
            return (d.strftime("%d.%m.%Y"), _DK_DAYS[d.weekday()])
        except ValueError:
            continue
    return (date_label, "")


def _render_error_xlsx(reason: str) -> bytes:
    """Minimal fallback when render fails. Owner gets an informative
    workbook instead of a 500."""
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Error"
        ws["A1"] = "BonBox: Excel generation failed"
        ws["A2"] = reason[:200]
        ws["A3"] = "Try again or send the close as text from the share sheet."
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except Exception:  # noqa: BLE001
        # Empty zipfile-shaped bytes so the response is at least valid
        return b"PK\x05\x06" + b"\x00" * 18
