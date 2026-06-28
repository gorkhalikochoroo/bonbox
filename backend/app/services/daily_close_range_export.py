"""Daily close range export — multi-day PDF + CSV + XLSX for accountant handoff.

Distinct from the per-close PDF (services/kasserapport_pdf.py) which
makes one polished receipt per closing day. This module aggregates a
DATE RANGE into a single document, with totals + averages so the
accountant can review a week / month / quarter at a glance.

Three formats:
  • PDF   — table with one row per close, totals + readiness badge.
            Danish labels for DKK tenants (matches the per-close PDF).
  • CSV   — same data tabular, semicolon-delimited + UTF-8 BOM so
            Danish-locale Excel opens it cleanly.
  • XLSX  — proper Excel workbook: typed columns (numbers as numbers,
            dates as dates), frozen header, auto-widths, Danish number
            format. The accountant can sort / filter / pivot directly.

Inputs are pre-filtered DailyClose ORM rows (date+user already
narrowed by the router); this service only formats.
"""
from __future__ import annotations

import csv
import io
from datetime import date

from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.daily_close import DailyClose, decode_breakdown
from app.utils.csv_safe import csv_safe


# ─── CSV ──────────────────────────────────────────────────────────────

_CSV_COLUMNS = [
    "date", "branch_id", "status",
    "revenue_total", "revenue_ex_moms", "moms_total", "moms_mode",
    "payment_total", "cash_expected", "cash_counted", "cash_difference",
    "tips_total", "tips_staff_count", "tips_per_person",
    "revenue_breakdown",   # encoded "food:12400|drinks:5800"
    "payment_breakdown",   # encoded "cash:4200|card:13500"
    "closed_by", "closed_at",
    "unlock_reason", "unlocked_by", "unlocked_at",
    "notes", "receipt_photo",
]


def _opt(v):
    """Format an optional float as fixed 2-decimal or empty string."""
    if v is None:
        return ""
    return f"{float(v):.2f}"


def closes_to_csv_bytes(closes: list[DailyClose]) -> bytes:
    """Serialize a list of DailyClose to CSV. UTF-8 BOM + semicolon
    delimiter so Danish Excel opens it cleanly with Æ/Ø/Å intact."""
    buf = io.StringIO()
    buf.write("﻿")
    writer = csv.writer(buf, delimiter=";")
    writer.writerow(_CSV_COLUMNS)
    for c in closes:
        writer.writerow([
            c.date.isoformat() if c.date else "",
            str(c.branch_id) if c.branch_id else "",
            csv_safe(getattr(c, "status", "") or ""),
            _opt(c.revenue_total),
            _opt(c.revenue_ex_moms),
            _opt(c.moms_total),
            csv_safe(getattr(c, "moms_mode", "") or ""),
            _opt(c.payment_total),
            _opt(c.cash_expected),
            _opt(c.cash_counted),
            _opt(c.cash_difference),
            _opt(c.tips_total),
            c.tips_staff_count if c.tips_staff_count is not None else "",
            _opt(c.tips_per_person),
            csv_safe(c.revenue_categories or ""),
            csv_safe(c.payment_categories or ""),
            csv_safe(c.closed_by or ""),
            c.closed_at.isoformat() if c.closed_at else "",
            csv_safe(getattr(c, "unlock_reason", "") or ""),
            csv_safe(getattr(c, "unlocked_by", "") or ""),
            (c.unlocked_at.isoformat() if getattr(c, "unlocked_at", None) else ""),
            csv_safe(c.notes or ""),
            csv_safe(getattr(c, "receipt_photo", "") or ""),
        ])
    return buf.getvalue().encode("utf-8")


# ─── Helpers shared by PDF + XLSX ─────────────────────────────────────

# Payment-method buckets we surface as their own columns in PDF + XLSX.
# Anything unrecognized rolls into "other" so the totals still tie out.
_PAY_BUCKETS = ["cash", "card", "mobilepay", "gift_card", "bank_transfer"]

# Card-brand / scheme splits are a BREAKDOWN of the `card` line, not extra
# methods — they ride along in payment_categories for accountant fidelity but
# summing them would double-count the card total. Mirrors the write-side
# invariant `_CARD_BREAKDOWN_KEYS` in routers/daily_close.py (the close UI
# shows brands UNDER the card line, never adding). Excluded from every bucket
# so _payments_sum() ties out to revenue for a normal terminal close.
_CARD_BRAND_KEYS = {"dankort", "visa", "mastercard", "softpay", "betalingskort"}


def _bucketed_payments(c: DailyClose) -> dict:
    """Decode encoded payment_categories into known buckets + 'other'.

    Card-brand scheme keys (dankort/visa/…) are dropped — they are a split of
    the `card` line, so counting them would inflate payments past revenue and
    false-flag reconciliation on the most common DK terminal close."""
    out = {k: 0.0 for k in _PAY_BUCKETS}
    out["other"] = 0.0
    raw = decode_breakdown(c.payment_categories) or {}
    for k, v in raw.items():
        try:
            amt = float(v or 0)
        except (TypeError, ValueError):
            continue
        norm = (k or "").lower().replace(" ", "_").replace("-", "_")
        if norm in _CARD_BRAND_KEYS:
            continue  # brand split of the card line — never add it on top
        # Danish ⟷ English payment-method synonyms. The write side persists
        # English keys (cash/card/mobilepay — see routers/daily_close.py:1565
        # kontant→cash, dankort→card; defensive "kontant" in pb at :893), but
        # manually-edited / legacy / imported closes can still carry Danish
        # spellings. Without this map kontant/kort fell into "other" and
        # vanished from the table's Kontant/Kort columns while still summing
        # into _payments_sum — so the reconciliation badge read "✓ afstemt"
        # above a table that visibly did NOT tie out. Normalize so every
        # spelling lands in the right visible column.
        norm = {
            "kontant": "cash",
            "kort": "card",
            # NB: "betalingskort" is intentionally NOT mapped here — it is a
            # card-brand split (in _CARD_BRAND_KEYS above) and is dropped
            # before reaching this map, never added on top of the card line.
            "mobile_pay": "mobilepay",
            "mobilepay_total": "mobilepay",
            "gavekort": "gift_card",
            "giftcard": "gift_card",
            "gift_cards": "gift_card",
            "bankoverforsel": "bank_transfer",
            "bankoverførsel": "bank_transfer",
            "bank_overforsel": "bank_transfer",
            "overforsel": "bank_transfer",
        }.get(norm, norm)
        if norm in _PAY_BUCKETS:
            out[norm] += amt
        else:
            out["other"] += amt
    return out


def _payments_sum(c: DailyClose) -> float:
    """Total of EVERY decoded payment bucket for a close — including the ones
    the PDF's narrow column set hides (gift_card / bank_transfer / other).

    This is the number that must tie out to ``revenue_total`` for a close to be
    bookkeeping-ready: every krone of revenue is collected by some method, so
    the methods must sum to the revenue. The per-close PDF reconciliation and
    the readiness badge both key on it."""
    return sum(_bucketed_payments(c).values())


# Per-row tie-out tolerance: a single close row whose decoded payments differ
# from its revenue by more than this (in kr) is FLAGGED inline in both the PDF
# and the XLSX — drafts included. 0.5 kr absorbs øre rounding but catches the
# real "Kort 1.850 vs Omsætning 1.070" (780 kr) mismatch the export must never
# present silently. Mirrors the aggregate RECON_TOL used below.
ROW_TIE_OUT_TOL = 0.50


def _row_ties_out(c: DailyClose) -> bool:
    """True when this close's decoded payments tie out to its revenue within
    ROW_TIE_OUT_TOL — OR when there are no reconcilable payment methods at all
    (a revenue-only / scan-and-lock close has nothing to tie out, so it is not
    flagged). False means: payments were recorded AND they do not match revenue
    → the row must carry the inline mismatch flag."""
    if not _has_reconcilable_payments(c):
        return True
    return abs(_payments_sum(c) - float(c.revenue_total or 0)) <= ROW_TIE_OUT_TOL


def _has_reconcilable_payments(c: DailyClose) -> bool:
    """True only when the close actually recorded payment-method amounts that
    sum to something. A revenue-only close (scan-and-lock / Z-report bottom
    line with no method split) or one carrying only card-brand splits has
    NOTHING to tie out — reconciling it would falsely flag the whole day's
    revenue as an 'Afvigelse' and wrongly drop a book-ready close to review."""
    return _payments_sum(c) > 0.005


def compose_business_address(profile) -> str:
    """One-line address that never double-prints the city.

    The stored ``address`` field very often ALREADY ends with the postal code +
    city (DK convention: "Carl Th. Dreyers Vej 244, 4. 3., 2500 Valby") while
    ``zipcode`` and ``city`` are ALSO populated separately. Naively joining
    ``address`` + "zip city" produced the "…, 2500 Valby, 2500 Valby" defect.

    Rule: append the "zip city" tail ONLY when the address does not already
    carry it — i.e. the full "zip city" is not a substring, AND we don't see
    both the zip and the city already inside the address (and, when no zip is
    set, the bare city is not already there). Conservative so a street that
    merely contains a city name is not wrongly suppressed."""
    addr = (getattr(profile, "address", None) or "").strip().rstrip(",").strip()
    zipc = (getattr(profile, "zipcode", None) or "").strip()
    city = (getattr(profile, "city", None) or "").strip()
    zipcity = " ".join(p for p in (zipc, city) if p).strip()
    if not addr:
        return zipcity
    al = addr.lower()
    already = (
        (bool(zipcity) and zipcity.lower() in al)
        or (bool(zipc) and bool(city) and zipc.lower() in al and city.lower() in al)
        or (bool(city) and not zipc and city.lower() in al)
    )
    if zipcity and not already:
        return f"{addr}, {zipcity}"
    return addr


def _fmt_kr(v, currency: str = "DKK") -> str:
    """Render a money value as the PDF's canonical Danish string ("1.850,00
    kr."). Wraps the gold money_dk formatter (same one the PDF/MOMS artifacts
    use) so the XLSX tie-out note text matches the PDF byte-for-byte. Imported
    lazily — keeps this module import-light and mirrors the PDF builder, which
    also imports money_dk inside the function."""
    from app.services.bonbox_pdf_kit import money_dk
    return money_dk(v, currency)


def _voucher_ranges(db: Session, user_id, dt: date) -> tuple[str, str]:
    """Return (sales_label, expense_label) bilagsnummer range for a day.

    Empty strings if no vouchers were recorded for that date. Used in
    both PDF and XLSX so the accountant can cross-check each row
    against the underlying voucher numbers."""
    if db is None or dt is None:
        return ("", "")
    try:
        from app.models.sale import Sale
        from app.models.expense import Expense
        smin, smax = (
            db.query(func.min(Sale.voucher_number), func.max(Sale.voucher_number))
            .filter(
                Sale.user_id == user_id,
                Sale.date == dt,
                Sale.voucher_number.is_not(None),
                Sale.is_deleted.isnot(True),
            )
            .one_or_none() or (None, None)
        )
        emin, emax = (
            db.query(func.min(Expense.voucher_number), func.max(Expense.voucher_number))
            .filter(
                Expense.user_id == user_id,
                Expense.date == dt,
                Expense.voucher_number.is_not(None),
                Expense.is_deleted.isnot(True),
            )
            .one_or_none() or (None, None)
        )
    except Exception:
        return ("", "")

    def _fmt(prefix, lo, hi, year):
        if lo is None:
            return ""
        if lo == hi:
            return f"{prefix}-{year}-{lo:04d}"
        return f"{prefix}-{year}-{lo:04d} → {prefix}-{year}-{hi:04d}"

    return (
        _fmt("S", smin, smax, dt.year),
        _fmt("E", emin, emax, dt.year),
    )


# ─── PDF ──────────────────────────────────────────────────────────────

def build_daily_close_range_pdf(
    closes: list[DailyClose],
    *,
    from_date: date,
    to_date: date,
    business_name: str = "",
    currency: str = "DKK",
    profile=None,
    db: Session | None = None,
    user_id=None,
    bilagsnummer: str = "",
) -> bytes:
    """Build a multi-day daily-close PDF report — accountant-grade.

    What this PDF gives the accountant:
      • Business identity header (company_name, CVR, address, period)
      • One row per close: Date, Bilag, Revenue, MOMS (25%), Net,
                            Cash, Card, MobilePay, Cash diff, Status
      • Totals row (excludes Draft closes — only confirmed closes count)
      • Readiness badge: "X af Y closes klar til bogføring"
      • Footer cites Bogføringsloven §10 + Momsbekendtgørelsen §57

    Localized for DKK (Danish labels), else English.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph, Spacer, Table, TableStyle, HRFlowable,
    )

    from app.services.bonbox_pdf_kit import money_dk, render_with_doc_hash
    from app.utils.document_hash import get_software_identifier
    from app.utils.time import utc_now

    DA = (currency == "DKK")

    L = {
        "title":       "Kasserapport" if DA else "Daily Close",
        "period":      "Periode" if DA else "Period",
        "closes":      "lukninger" if DA else "closes",
        "close_one":   "lukning" if DA else "close",
        "across_days": "over" if DA else "across",
        "days":        "dage" if DA else "days",
        "day_one":     "dag" if DA else "day",
        # Header summary — count confirmed closes and drafts separately so a
        # single Kladde never reads as "1 lukning" (which implies a booked,
        # confirmed close). "bekræftede" + "kladde(r)" keep the distinction
        # honest for the revisor.
        "confirmed_word":   "bekræftede" if DA else "confirmed",
        "draft_word":       "kladde" if DA else "draft",
        "drafts_word":      "kladder" if DA else "drafts",
        "amounts_in":  "Alle beløb i" if DA else "All amounts in",
        "no_closes":   "Ingen lukninger i denne periode." if DA else "No daily closes in this range.",
        # KPI band labels (kept short for the 4-up grid)
        "kpi_rev":     "Omsætning" if DA else "Revenue",
        "kpi_moms":    "Salgsmoms" if DA else "Output VAT",
        "kpi_net":     "Netto (uden moms)" if DA else "Net (excl. VAT)",
        "kpi_tips":    "Drikkepenge" if DA else "Tips",
        # Table headers
        "h_date":      "Dato" if DA else "Date",
        "h_bilag":     "Bilag" if DA else "Voucher",
        "h_rev":       "Omsætning" if DA else "Revenue",
        "h_moms":      "Moms 25%" if DA else "VAT 25%",
        "h_net":       "Netto" if DA else "Net",
        "h_cash":      "Kontant" if DA else "Cash",
        "h_card":      "Kort" if DA else "Card",
        "h_mobilepay": "MobilePay" if DA else "MobilePay",
        "h_other":     "Øvrige" if DA else "Other",
        "h_diff":      "Kassediff." if DA else "Diff",
        "h_status":    "Status" if DA else "Status",
        # Status badges
        "locked":      "Lukket" if DA else "Confirmed",
        "draft":       "Kladde" if DA else "Draft",
        # Per-row tie-out footnote — explains the "!" marker that flags any row
        # whose payments do not equal its revenue (drafts included). The "!" is
        # rendered as a bold amber glyph; this text is the explanation after it.
        # The kasserapport must never present a non-tying row silently.
        "row_flag_note": (
            "betalinger stemmer ikke med omsætning"
            if DA else
            "payments do not match revenue"
        ),
        # Totals row
        "totals":      "I alt (bekræftede)" if DA else "Totals (confirmed)",
        "draft_excl":  "Kladder ikke medregnet" if DA else "Drafts not summed",
        # Readiness
        "ready":       "klar til bogføring" if DA else "ready for booking",
        "review":      "skal gennemgås" if DA else "need review",
        # ±100 kr cash-drawer tolerance — stated openly so the threshold the
        # readiness badge applies (abs(cash_difference) <= 100) is never silent.
        "ready_tol":   ("kassedifference inden for ±100 kr." if DA
                        else "cash difference within ±100 kr."),
        # Kasseafstemning (cash-drawer reconciliation: forventet − optalt = diff)
        "kasse_title":  "Kasseafstemning" if DA else "Cash reconciliation",
        "kasse_exp":    "forventet" if DA else "expected",
        "kasse_cnt":    "optalt" if DA else "counted",
        # Document voucher number (header) — matches the gold MOMS-PDF label
        "bilag_label": "Bilagsnr" if DA else "Voucher",
        # Payments ↔ revenue reconciliation (accountant-grade)
        "recon_ok":    "Betalinger afstemt med omsætning" if DA else "Payments reconciled to revenue",
        "recon_off":   "Betalinger stemmer ikke med omsætning" if DA else "Payments do not reconcile to revenue",
        "recon_pay":   "Betalinger" if DA else "Payments",
        "recon_rev":   "Omsætning" if DA else "Revenue",
        "recon_delta": "Afvigelse" if DA else "Difference",
        "recon_review": ("lukning(er) skal gennemgås før bogføring"
                         if DA else "close(s) need review before booking"),
        # Signature line (two-person control, Bogføringsloven §10)
        "sig_counted":  "Optalt af" if DA else "Counted by",
        "sig_approved": "Godkendt af" if DA else "Approved by",
        "sig_date":     "Dato" if DA else "Date",
        # Footer
        "footer_law":  (
            "Genereret af BonBox · Opbevares iht. Bogføringsloven §10 (5 år). "
            "Salgsmoms beregnet pr. Momsbekendtgørelsen §57."
        ) if DA else (
            "Generated by BonBox · Retention per Bogføringsloven §10 (5 years). "
            "VAT computed per Momsbekendtgørelsen §57."
        ),
    }

    # Sort ascending so the report reads chronologically.
    closes_sorted = sorted(closes, key=lambda c: c.date or date.min)

    # Danish-style date for header period
    if DA:
        _MM = ["jan", "feb", "mar", "apr", "maj", "jun",
               "jul", "aug", "sep", "okt", "nov", "dec"]
        def _date_short(d):
            return f"{d.day}. {_MM[d.month - 1]} {d.year}" if d else ""
    else:
        def _date_short(d):
            return d.strftime("%d %b %Y") if d else ""

    # ─── Provenance footer inputs (computed ONCE, before rendering) ──────
    # All fail-soft: footer derivation must never break the PDF.
    software_id = get_software_identifier()
    try:
        generated_at_str = utc_now().strftime("%Y-%m-%d %H:%M UTC")
    except Exception:  # noqa: BLE001 — clock/strftime should never crash the PDF
        generated_at_str = ""
    generator_email = ""
    if db is not None and user_id is not None:
        try:
            from app.models.user import User
            _u = db.query(User).filter(User.id == user_id).first()
            generator_email = (getattr(_u, "email", None) or "") if _u else ""
        except Exception:  # noqa: BLE001 — never let footer derivation break the PDF
            generator_email = ""

    INK = colors.HexColor("#171717")
    MUTED = colors.HexColor("#6b7280")
    DIVIDER = colors.HexColor("#e5e7eb")
    EMERALD = colors.HexColor("#065f46")
    AMBER = colors.HexColor("#92400e")
    EMERALD_BG = colors.HexColor("#d1fae5")
    AMBER_BG = colors.HexColor("#fef3c7")

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=15,
                        textColor=INK, fontName="Helvetica-Bold",
                        leading=18, spaceAfter=2)
    subtitle = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=9,
                              textColor=MUTED, leading=12, spaceAfter=6)
    section = ParagraphStyle("Sect", parent=styles["Normal"], fontSize=8,
                             textColor=MUTED, fontName="Helvetica-Bold",
                             leading=11, spaceBefore=8, spaceAfter=2)
    note = ParagraphStyle("Note", parent=styles["Normal"], fontSize=8,
                          textColor=MUTED, fontName="Helvetica-Oblique",
                          leading=11)
    # Centered cell style for the Status column when it carries the inline "!"
    # tie-out marker (a Paragraph, not a bare string, so the "!" can be amber).
    status_style = ParagraphStyle("Status", parent=styles["Normal"], fontSize=8.5,
                                  textColor=INK, fontName="Helvetica",
                                  leading=10, alignment=1)  # 1 = TA_CENTER

    def _make_story():
        # Returns a FRESH list of flowables each call — render_with_doc_hash
        # invokes this once per pass (reportlab mutates flowables in-place, so
        # pass 1 and pass 2 each need brand-new objects). All numeric data is
        # plain-value closure capture; only the Paragraph/Table/Spacer/
        # HRFlowable OBJECTS are created fresh here.
        story = []

        # ─── Business identity header ────────────────────────────────────
        biz_line = f"<font name='Helvetica-Bold' size='12'>{business_name or '—'}</font>"
        biz_meta = []
        if profile:
            if getattr(profile, "org_number", None):
                biz_meta.append(f"CVR {profile.org_number}")
            addr_line = compose_business_address(profile)
            if addr_line:
                biz_meta.append(addr_line)
        head_left = biz_line
        if biz_meta:
            head_left += f"<br/><font color='#6b7280' size='9'>{' · '.join(biz_meta)}</font>"

        head_right = (
            f"<font name='Helvetica-Bold' size='11'>{L['title']}</font>"
            f"<br/><font color='#6b7280' size='9'>"
            f"{_date_short(from_date)} → {_date_short(to_date)}"
            f"</font>"
        )
        # Document voucher number — gives the revisor a stable per-document
        # reference (mirrors build_moms_filing_pdf). Only shown when supplied.
        if bilagsnummer:
            head_right += (
                f"<br/><font color='#6b7280' size='8'>"
                f"{L['bilag_label']} {bilagsnummer}</font>"
            )

        head_table = Table(
            [[Paragraph(head_left, subtitle), Paragraph(head_right, subtitle)]],
            colWidths=[160*mm, 110*mm],
        )
        head_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ]))
        story.append(head_table)
        story.append(HRFlowable(width="100%", thickness=0.5, color=DIVIDER, spaceBefore=2, spaceAfter=8))

        confirmed = [c for c in closes_sorted if (getattr(c, "status", None) or "confirmed") == "confirmed"]
        drafts = [c for c in closes_sorted if (getattr(c, "status", None) or "confirmed") != "confirmed"]
        n_conf = len(confirmed)
        span_days = (to_date - from_date).days + 1 if to_date and from_date and to_date >= from_date else 1

        # Singular/plural agreement on confirmed + drafts SEPARATELY so a lone
        # Kladde never reads as "1 lukning" (which implies a booked close). E.g.
        # "0 bekræftede · 1 kladde over 30 dage" vs "3 bekræftede · 2 kladder
        # over 7 dage". n_draft uses the plural-aware draft words.
        n_draft = len(drafts)
        days_word = L["day_one"] if span_days == 1 else L["days"]
        confirmed_part = f"{n_conf} {L['confirmed_word']}"
        draft_word = L["draft_word"] if n_draft == 1 else L["drafts_word"]
        draft_part = f"{n_draft} {draft_word}"
        story.append(Paragraph(
            f"{confirmed_part} · {draft_part} {L['across_days']} {span_days} {days_word}"
            f"  ·  {L['amounts_in']} {currency}",
            subtitle,
        ))

        if not closes_sorted:
            story.append(Spacer(1, 6))
            story.append(Paragraph(L["no_closes"], note))
            story.append(Spacer(1, 12))
            story.append(Paragraph(L["footer_law"], note))
            return story

        # ─── KPI band (only sums CONFIRMED — drafts excluded so totals are
        #   honest for the accountant) ────────────────────────────────────
        def _sum(attr):
            return sum(float(getattr(c, attr, 0) or 0) for c in confirmed)

        total_revenue = _sum("revenue_total")
        total_moms = _sum("moms_total")
        total_net = sum(
            float(c.revenue_ex_moms or 0) if c.revenue_ex_moms is not None
            else max(float(c.revenue_total or 0) - float(c.moms_total or 0), 0)
            for c in confirmed
        )
        total_tips = _sum("tips_total")

        def _fmt(v):
            # Canonical money formatter — the ONE every BonBox export must use
            # (mirrors the gold MOMS-PDF + the per-close kasserapport_pdf). DKK
            # → "1.234,56 kr." (period thousands, comma decimal, "kr." unit);
            # None / "" / non-numeric → "—". Previously this builder rolled its
            # own formatter that produced Danish digits but DROPPED the "kr."
            # unit, so the range PDF disagreed with every other artifact and a
            # revisor saw bare "15.000,00". money_dk also snaps negative-zero
            # dust to "0,00 kr." for free.
            return money_dk(v, currency)

        kpi_rows = [[
            Paragraph(f"<font color='#6b7280' size='8'>{L['kpi_rev']}</font><br/><font name='Helvetica-Bold' size='13'>{_fmt(total_revenue)}</font>", subtitle),
            Paragraph(f"<font color='#6b7280' size='8'>{L['kpi_moms']}</font><br/><font name='Helvetica-Bold' size='13'>{_fmt(total_moms)}</font>", subtitle),
            Paragraph(f"<font color='#6b7280' size='8'>{L['kpi_net']}</font><br/><font name='Helvetica-Bold' size='13'>{_fmt(total_net)}</font>", subtitle),
            Paragraph(f"<font color='#6b7280' size='8'>{L['kpi_tips']}</font><br/><font name='Helvetica-Bold' size='13'>{_fmt(total_tips)}</font>", subtitle),
        ]]
        kpi = Table(kpi_rows, colWidths=[68*mm, 68*mm, 68*mm, 66*mm])
        kpi.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f9fafb")),
            ("BOX", (0, 0), (-1, -1), 0.5, DIVIDER),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, DIVIDER),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ]))
        story.append(kpi)
        story.append(Spacer(1, 8))

        # ─── Per-close table — accountant columns ────────────────────────
        # "Øvrige" carries every payment method outside the named columns
        # (gift_card / bank_transfer / unrecognized) so the row's payment
        # cells ALWAYS visibly sum to omsætning. Without it, money in those
        # buckets was invisible on the page yet counted in the reconciliation
        # — a "✓ afstemt" badge above a table that didn't add up.
        table_data = [[
            L["h_date"], L["h_bilag"], L["h_rev"], L["h_moms"], L["h_net"],
            L["h_cash"], L["h_card"], L["h_mobilepay"], L["h_other"],
            L["h_diff"], L["h_status"],
        ]]

        # Track whether ANY row (drafts included) failed its per-row tie-out so
        # we can emit the footnote that explains the inline "!" marker.
        any_row_flagged = False

        for c in closes_sorted:
            pay = _bucketed_payments(c)
            vsales, _vexp = _voucher_ranges(db, user_id, c.date)
            revenue = float(c.revenue_total or 0)
            moms = float(c.moms_total or 0) if c.moms_total is not None else None
            net = (
                float(c.revenue_ex_moms or 0) if c.revenue_ex_moms is not None
                else (revenue - (moms or 0))
            )
            cash_diff = c.cash_difference
            cash_diff_str = ""
            if cash_diff is not None:
                sign = "+" if cash_diff > 0 else ""
                cash_diff_str = f"{sign}{_fmt(cash_diff)}"
            status = (getattr(c, "status", None) or "confirmed")
            # Residual = every bucket outside the three named columns, so the
            # visible payment cells (Kontant+Kort+MobilePay+Øvrige) always add
            # up to the day's collected payments.
            other_amt = pay["gift_card"] + pay["bank_transfer"] + pay["other"]
            # Per-row tie-out flag — DRAFTS INCLUDED. A row whose recorded
            # payments do not equal its revenue carries an amber "!" next to its
            # status badge (typographic marker, no emoji — matches the badge
            # marks elsewhere). The kasserapport never shows a non-tying row,
            # even a Kladde, without saying so.
            status_label = L["locked"] if status == "confirmed" else L["draft"]
            if not _row_ties_out(c):
                any_row_flagged = True
                status_cell = Paragraph(
                    f"<font name='Helvetica-Bold' color='{AMBER.hexval()}'>!</font>"
                    f" {status_label}",
                    status_style,
                )
            else:
                status_cell = status_label
            table_data.append([
                _date_short(c.date),
                vsales or "—",
                _fmt(revenue),
                _fmt(moms) if moms is not None else "—",
                _fmt(net),
                _fmt(pay["cash"]) if pay["cash"] else "—",
                _fmt(pay["card"]) if pay["card"] else "—",
                _fmt(pay["mobilepay"]) if pay["mobilepay"] else "—",
                _fmt(other_amt) if other_amt else "—",
                cash_diff_str or "—",
                status_cell,
            ])

        # Totals (confirmed only)
        sum_cash = sum(_bucketed_payments(c)["cash"] for c in confirmed)
        sum_card = sum(_bucketed_payments(c)["card"] for c in confirmed)
        sum_mp = sum(_bucketed_payments(c)["mobilepay"] for c in confirmed)
        sum_other = sum(
            _bucketed_payments(c)["gift_card"]
            + _bucketed_payments(c)["bank_transfer"]
            + _bucketed_payments(c)["other"]
            for c in confirmed
        )
        sum_diff = sum(
            float(c.cash_difference or 0) for c in confirmed
            if c.cash_difference is not None
        )

        table_data.append([
            f"{L['totals']} ({n_conf})", "",
            _fmt(total_revenue), _fmt(total_moms), _fmt(total_net),
            _fmt(sum_cash) if sum_cash else "—",
            _fmt(sum_card) if sum_card else "—",
            _fmt(sum_mp) if sum_mp else "—",
            _fmt(sum_other) if sum_other else "—",
            _fmt(sum_diff) if any(c.cash_difference is not None for c in confirmed) else "—",
            "",
        ])

        col_widths = [25*mm, 25*mm, 24*mm, 21*mm, 23*mm, 22*mm, 21*mm, 23*mm, 21*mm, 21*mm, 18*mm]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 8.5),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
            ("TEXTCOLOR", (0, 0), (-1, 0), MUTED),
            ("ALIGN", (2, 0), (9, -1), "RIGHT"),
            ("ALIGN", (10, 0), (10, -1), "CENTER"),
            ("ALIGN", (1, 0), (1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ("TOPPADDING", (0, 0), (-1, 0), 6),
            ("BOTTOMPADDING", (0, 1), (-1, -2), 5),
            ("TOPPADDING", (0, 1), (-1, -2), 5),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, DIVIDER),
            ("LINEBELOW", (0, 1), (-1, -2), 0.25, DIVIDER),
            ("LINEABOVE", (0, -1), (-1, -1), 1, EMERALD),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), EMERALD_BG),
            ("TEXTCOLOR", (0, -1), (-1, -1), EMERALD),
        ]))
        story.append(table)

        # ─── Per-row tie-out footnote (if any row was flagged) ───────────
        # Explains the amber "!" marker carried by any row — drafts included —
        # whose payments do not equal its revenue. Placed directly under the
        # table so the marker and its meaning sit together.
        if any_row_flagged:
            story.append(Spacer(1, 4))
            story.append(Paragraph(
                f"<font color='{AMBER.hexval()}' size='8.5'>"
                f"<font name='Helvetica-Bold'>!</font>"
                f" {L['row_flag_note']}</font>",
                note,
            ))

        # ─── Draft note (if any) ─────────────────────────────────────────
        if drafts:
            story.append(Spacer(1, 4))
            story.append(Paragraph(
                f"<font color='#92400e' size='8.5'>"
                f"<font name='Helvetica-Bold'>!</font> {len(drafts)} "
                f"{L['draft'].lower()} — {L['draft_excl']}.</font>",
                note,
            ))

        # ─── Kasseafstemning (drawer reconciliation) ─────────────────────
        # The "Kassediff." column shows only the NET difference; a revisor
        # needs the derivation (forventet − optalt = difference) to trust the
        # number. We surface it as one quiet line per close that actually
        # counted its drawer, with the sign on the difference. forventet =
        # register/expected, optalt = counted (read for XLSX but previously
        # never shown on the PDF). Confirmed closes only.
        counted_closes = [
            c for c in confirmed
            if c.cash_counted is not None and c.cash_expected is not None
        ]
        if counted_closes:
            story.append(Spacer(1, 6))
            story.append(Paragraph(
                f"<font color='#6b7280' name='Helvetica-Bold' size='8'>"
                f"{L['kasse_title']}</font>",
                note,
            ))
            for c in counted_closes:
                exp = float(c.cash_expected or 0)
                cnt = float(c.cash_counted or 0)
                diff = (float(c.cash_difference) if c.cash_difference is not None
                        else round(cnt - exp, 2))
                diff_sign = "+" if diff > 0 else ""
                # Emerald when within ±100 kr, amber when outside — status
                # colour only, no rainbow.
                diff_col = EMERALD if abs(diff) <= 100 else AMBER
                story.append(Paragraph(
                    f"<font color='#6b7280' size='8'>"
                    f"{_date_short(c.date)} — {L['kasse_exp']} {_fmt(exp)} − "
                    f"{L['kasse_cnt']} {_fmt(cnt)} = "
                    f"<font color='{diff_col.hexval()}' name='Helvetica-Bold'>"
                    f"{diff_sign}{_fmt(diff)}</font></font>",
                    note,
                ))

        # ─── Payments ↔ revenue reconciliation (accountant-grade) ────────
        # A kasserapport must TIE OUT: every krone of revenue is collected by
        # some payment method, so the methods must sum to the revenue. We
        # reconcile on CONFIRMED closes (drafts are excluded from totals) and
        # surface any close that does not balance — silently presenting
        # non-tying numbers to a revisor is exactly what "accountant-grade"
        # exists to prevent. RECON_TOL absorbs øre rounding but catches real
        # gaps (the per-close payment sum includes hidden buckets like
        # gift_card / bank_transfer / other that the narrow table omits).
        RECON_TOL = 0.50  # 50 øre
        # Only closes that actually recorded payment methods can be reconciled —
        # a revenue-only close has nothing to tie out (see _has_reconcilable_).
        recon_closes = [c for c in confirmed if _has_reconcilable_payments(c)]
        unbalanced = [
            c for c in recon_closes
            if abs(_payments_sum(c) - float(c.revenue_total or 0)) > RECON_TOL
        ]
        if recon_closes:
            story.append(Spacer(1, 6))
            if unbalanced:
                # Sum of ABSOLUTE per-close deviations — never 0 while a close
                # is off (a NET figure could cancel out and read "0,00" beside
                # a "needs review" warning, which is self-contradictory).
                abs_dev = sum(
                    abs(_payments_sum(c) - float(c.revenue_total or 0))
                    for c in unbalanced
                )
                story.append(Paragraph(
                    f"<font color='{AMBER.hexval()}' size='8.5'>"
                    f"<font name='Helvetica-Bold'>!</font> {L['recon_off']}: "
                    f"{len(unbalanced)} {L['recon_review']} "
                    f"({L['recon_delta']} {_fmt(abs_dev)}).</font>",
                    note,
                ))
            else:
                net = (
                    sum(_payments_sum(c) for c in recon_closes)
                    - sum(float(c.revenue_total or 0) for c in recon_closes)
                )
                story.append(Paragraph(
                    f"<font color='{EMERALD.hexval()}' size='8.5'>"
                    f"<font name='Helvetica-Bold'>•</font> {L['recon_ok']} "
                    f"({L['recon_delta']} {_fmt(net)}).</font>",
                    note,
                ))

        # ─── Readiness badge ─────────────────────────────────────────────
        # A close counts as "ready for booking" only when ALL three hold:
        # MOMS computed, cash drawer within ±100 (or not counted), AND
        # payments reconcile to revenue. The third condition is the honesty
        # fix — previously a close with a payment/revenue mismatch could still
        # claim "klar til bogføring".
        ready_count = sum(
            1 for c in confirmed
            if (c.moms_total is not None)
            and (c.cash_counted is None or abs(float(c.cash_difference or 0)) <= 100)
            and (not _has_reconcilable_payments(c)
                 or abs(_payments_sum(c) - float(c.revenue_total or 0)) <= RECON_TOL)
        )
        all_ready = (ready_count == n_conf and n_conf > 0)
        badge_text = L["ready"] if all_ready else L["review"]
        badge_color = EMERALD if all_ready else AMBER
        badge_bg = EMERALD_BG if all_ready else AMBER_BG
        # Neutral typographic status mark (no emoji — design lock): emerald "•"
        # when ready, amber "!" when review is needed. Colour carries the
        # status, not a coloured glyph.
        icon = "•" if all_ready else "!"

        story.append(Spacer(1, 8))
        badge = Table(
            [[Paragraph(
                f"<font name='Helvetica-Bold' color='{badge_color.hexval()}' size='10'>"
                f"{icon} {ready_count} / {n_conf} {badge_text}</font>"
                f"<font color='#6b7280' size='8'> · {L['ready_tol']}</font>",
                subtitle,
            )]],
            colWidths=[270*mm],
        )
        badge.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), badge_bg),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(badge)

        # ─── Signature line (two-person control, Bogføringsloven §10) ─────
        # The revisor expects a signed cash report: who counted the drawer +
        # who approved it for booking, with a date. Kept to a single quiet
        # line so it never crowds the landscape body.
        story.append(Spacer(1, 14))
        story.append(Paragraph(
            f"<font color='#6b7280' size='8'>"
            f"{L['sig_counted']}: ______________________&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;"
            f"{L['sig_approved']}: ______________________&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;"
            f"{L['sig_date']}: ____________</font>",
            note,
        ))

        # ─── Footer ──────────────────────────────────────────────────────
        story.append(Spacer(1, 6))
        story.append(HRFlowable(width="100%", thickness=0.5, color=DIVIDER, spaceBefore=2, spaceAfter=4))
        story.append(Paragraph(L["footer_law"], note))

        return story

    # ─── 2-pass render with the accountant-grade provenance footer ───────
    # Content-aligned margins (NOT symmetric 22mm): the wide landscape tables
    # are 270mm (head_table 160+110, KPI band, badge, per-close table). On
    # landscape A4 (297mm) a 271mm frame fits them with ~1mm spare and centres
    # them — 14mm left / 12mm right gives 297−14−12 = 271mm. Symmetric 22mm
    # would leave only 253mm and overrun the right margin by ~17mm. The footer
    # now anchors to these 14/12 content edges (not a hardcoded 22mm), so
    # Doc-hash / Side X/Y line up with the table edges. bottom_mm=20 keeps the
    # y=10mm footer clear of body content. render_with_doc_hash calls
    # _make_story once per pass (fresh flowables each time).
    period_subject = f"{from_date.isoformat()} → {to_date.isoformat()}"
    return render_with_doc_hash(
        _make_story,
        pagesize=landscape(A4),
        left_mm=14, right_mm=12, top_mm=14, bottom_mm=20,
        title=f"{L['title']} — {business_name}",
        author="BonBox",
        subject=period_subject,
        software_id=software_id,
        generated_at_str=generated_at_str,
        generator_email=generator_email,
        is_danish=DA,
    )


# ─── XLSX ─────────────────────────────────────────────────────────────

def build_daily_close_range_xlsx(
    closes: list[DailyClose],
    *,
    from_date: date,
    to_date: date,
    business_name: str = "",
    currency: str = "DKK",
    profile=None,
    db: Session | None = None,
    user_id=None,
) -> bytes:
    """Build an Excel workbook with two sheets:

      1. "Daily Close" — one row per close, typed columns (date as
         date, money as number with `# ##0,00` format). Frozen header,
         auto-widths, totals row using SUM() formulas so the accountant
         can edit rows and totals recompute live.
      2. "Summary" — KPI overview + business header.

    Accountants overwhelmingly prefer XLSX over PDF for end-of-period
    handoff because they can sort/filter/pivot. PDF stays available
    for the "send a one-pager to the bookkeeper" flow.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    DA = (currency == "DKK")

    # Localized headers — same vocabulary as the PDF for consistency
    H = (
        ["Dato", "Bilag (salg)", "Bilag (udgift)",
         "Omsætning", "Salgsmoms 25%", "Netto (uden moms)",
         "Kontant", "Kort", "MobilePay", "Gavekort", "Bank", "Andet",
         "Betalinger i alt", "Forventet kontant", "Optalt kontant", "Kasse-diff",
         "Drikkepenge", "Antal medarbejdere", "Pr. medarbejder",
         "Status", "Lukket af", "Lukket dato", "Bemærkninger"]
        if DA else
        ["Date", "Sales voucher", "Expense voucher",
         "Revenue", "Output VAT 25%", "Net (excl. VAT)",
         "Cash", "Card", "MobilePay", "Gift card", "Bank transfer", "Other",
         "Payments total", "Expected cash", "Counted cash", "Cash diff",
         "Tips", "Staff count", "Per person",
         "Status", "Closed by", "Closed at", "Notes"]
    )

    wb = Workbook()

    # ─── Sheet 1: Summary ─────────────────────────────────────────────
    s1 = wb.active
    s1.title = "Oversigt" if DA else "Summary"
    s1.column_dimensions["A"].width = 28
    s1.column_dimensions["B"].width = 40

    bold = Font(bold=True)
    big = Font(size=14, bold=True)

    s1["A1"] = business_name or "—"
    s1["A1"].font = big
    if profile:
        row = 2
        if getattr(profile, "org_number", None):
            s1.cell(row=row, column=1, value="CVR")
            s1.cell(row=row, column=2, value=str(profile.org_number))
            row += 1
        addr_line = compose_business_address(profile)
        if addr_line:
            s1.cell(row=row, column=1, value="Adresse" if DA else "Address")
            s1.cell(row=row, column=2, value=addr_line)
            row += 1

    period_row = 6
    s1.cell(row=period_row, column=1, value="Periode" if DA else "Period").font = bold
    s1.cell(row=period_row, column=2, value=f"{from_date.isoformat()} → {to_date.isoformat()}")
    s1.cell(row=period_row + 1, column=1, value="Valuta" if DA else "Currency").font = bold
    s1.cell(row=period_row + 1, column=2, value=currency)
    s1.cell(row=period_row + 2, column=1, value="Antal lukninger" if DA else "Close count").font = bold
    s1.cell(row=period_row + 2, column=2, value=len(closes))

    confirmed = [c for c in closes if (getattr(c, "status", None) or "confirmed") == "confirmed"]

    total_revenue = sum(float(c.revenue_total or 0) for c in confirmed)
    total_moms = sum(float(c.moms_total or 0) for c in confirmed)
    total_net = sum(
        float(c.revenue_ex_moms or 0) if c.revenue_ex_moms is not None
        else max(float(c.revenue_total or 0) - float(c.moms_total or 0), 0)
        for c in confirmed
    )
    total_tips = sum(float(c.tips_total or 0) for c in confirmed)

    # DKK money DISPLAY format — aligned to the PDF's "1.070,00 kr." (period
    # thousands, comma decimal, " kr." suffix). Cells stay NUMERIC (real
    # numbers the accountant can SUM); only the rendered string changes.
    #
    # We write the EXPLICIT Danish literal "#.##0,00" (period group / comma
    # decimal) rather than the US-grammar "#,##0.00". openpyxl stores the
    # format code VERBATIM in the sheet XML, and Excel/LibreOffice honour the
    # literal separators in a stored custom format regardless of the opening
    # machine's locale — so a US-locale accountant still sees Danish "1.070,00
    # kr.", matching the PDF. The trailing literal is quoted: " kr." (leading
    # space inside the quotes). Verified by reloading the saved file (see the
    # scratchpad harness): cell.number_format round-trips to this exact code.
    money_fmt = '#.##0,00" kr."' if DA else '#,##0.00" kr."'
    kpi_rows = [
        ("Omsætning i alt" if DA else "Total revenue", total_revenue),
        ("Salgsmoms i alt" if DA else "Total output VAT", total_moms),
        ("Netto (uden moms)" if DA else "Net (excl. VAT)", total_net),
        ("Drikkepenge i alt" if DA else "Total tips", total_tips),
    ]
    k_start = period_row + 4
    s1.cell(row=k_start, column=1, value="Totaler (kun bekræftede)" if DA else "Totals (confirmed only)").font = bold
    for i, (label, val) in enumerate(kpi_rows):
        r = k_start + 1 + i
        s1.cell(row=r, column=1, value=label)
        c = s1.cell(row=r, column=2, value=val)
        c.number_format = money_fmt
        c.alignment = Alignment(horizontal="right")

    # ─── Sheet 2: Daily Close detail ─────────────────────────────────
    s2 = wb.create_sheet("Kasserapport" if DA else "Daily Close")
    header_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    thin = Side(border_style="thin", color="E5E7EB")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for col_idx, label in enumerate(H, start=1):
        cell = s2.cell(row=1, column=col_idx, value=label)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "right", vertical="center")
        cell.border = border
    s2.freeze_panes = "A2"

    sorted_closes = sorted(closes, key=lambda c: c.date or date.min)
    for r, c in enumerate(sorted_closes, start=2):
        pay = _bucketed_payments(c)
        vsales, vexp = _voucher_ranges(db, user_id, c.date)
        revenue = float(c.revenue_total or 0) if c.revenue_total is not None else None
        moms = float(c.moms_total or 0) if c.moms_total is not None else None
        net = (
            float(c.revenue_ex_moms) if c.revenue_ex_moms is not None
            else (revenue - (moms or 0)) if revenue is not None else None
        )
        # Per-row tie-out note (Bemærkninger) — DRAFTS INCLUDED. When a close's
        # recorded payments don't equal its revenue, APPEND (never overwrite)
        # an explicit "Betalinger (X kr.) ≠ omsætning (Y kr.)" note to whatever
        # the owner already wrote, mirroring the PDF's "!" row flag. _fmt_kr
        # renders the same currency string as the PDF (DKK → "1.850,00 kr.").
        remarks = c.notes or ""
        if not _row_ties_out(c):
            pay_total = _payments_sum(c)
            rev = float(c.revenue_total or 0)
            mismatch = (
                f"Betalinger ({_fmt_kr(pay_total, currency)}) ≠ omsætning ({_fmt_kr(rev, currency)})"
                if DA else
                f"Payments ({_fmt_kr(pay_total, currency)}) ≠ revenue ({_fmt_kr(rev, currency)})"
            )
            remarks = f"{remarks} · {mismatch}" if remarks else mismatch
        row_values = [
            c.date, vsales, vexp,
            revenue, moms, net,
            pay["cash"] or None, pay["card"] or None, pay["mobilepay"] or None,
            pay["gift_card"] or None, pay["bank_transfer"] or None, pay["other"] or None,
            float(c.payment_total) if c.payment_total is not None else None,
            float(c.cash_expected) if c.cash_expected is not None else None,
            float(c.cash_counted) if c.cash_counted is not None else None,
            float(c.cash_difference) if c.cash_difference is not None else None,
            float(c.tips_total) if c.tips_total is not None else None,
            c.tips_staff_count,
            float(c.tips_per_person) if c.tips_per_person is not None else None,
            ("Lukket" if DA else "Confirmed") if (getattr(c, "status", None) or "confirmed") == "confirmed"
                else ("Kladde" if DA else "Draft"),
            c.closed_by or "",
            c.closed_at.replace(tzinfo=None) if c.closed_at else None,
            remarks,
        ]
        for col_idx, val in enumerate(row_values, start=1):
            cell = s2.cell(row=r, column=col_idx, value=val)
            cell.border = border
            if col_idx == 1 and val is not None:
                cell.number_format = "yyyy-mm-dd"
            elif col_idx >= 4 and col_idx <= 19 and isinstance(val, (int, float)):
                cell.number_format = money_fmt if col_idx != 18 else "0"  # staff count = integer
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 22 and val is not None:
                cell.number_format = "yyyy-mm-dd hh:mm"

    # Totals row with formulas (so the accountant can audit + edit)
    if sorted_closes:
        first = 2
        last = len(sorted_closes) + 1
        totals_row = last + 1
        s2.cell(row=totals_row, column=1, value=("I alt (bekræftede)" if DA else "Totals (confirmed)")).font = bold
        # Sum formulas for money columns (4..17)
        money_cols = list(range(4, 18))  # Revenue..Tips
        for col_idx in money_cols:
            letter = get_column_letter(col_idx)
            cell = s2.cell(row=totals_row, column=col_idx,
                           value=f"=SUM({letter}{first}:{letter}{last})")
            cell.font = bold
            cell.number_format = money_fmt
            cell.alignment = Alignment(horizontal="right")
            cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        for col_idx in [1, 2, 3, 18, 19, 20, 21, 22, 23]:
            cell = s2.cell(row=totals_row, column=col_idx)
            cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

    # Auto-ish column widths
    widths = [11, 16, 16, 13, 13, 14, 11, 11, 11, 11, 11, 11, 14, 13, 13, 12, 13, 11, 12, 12, 14, 18, 28]
    for i, w in enumerate(widths, start=1):
        s2.column_dimensions[get_column_letter(i)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
