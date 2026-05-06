"""Tests for the Mirabelle-format close .xlsx workbook.

Unlike PDF, openpyxl-generated workbooks ARE inspectable byte-for-byte
(they're just zipped XML), so we can:
  • Open the bytes back into a Workbook
  • Walk the cells
  • Verify labels, numeric values, number_formats, fills

This makes the Excel renderer easier to assert against than the PDF.
We still check the contract: never raises, valid xlsx, fail-closed
fallback on garbage input.
"""
from __future__ import annotations

import io

from openpyxl import load_workbook

from app.services.kasserapport_excel import (
    _danish_date_label,
    _money_value,
    render_close_xlsx,
)


# ─── Shared fixture: real Mirabelle Saturday 9.3 numbers ──────────────

def _abigail_aggregated() -> dict:
    """Real Mirabelle Saturday 9.3 numbers from the photographed Excel."""
    return {
        "closed_by": "Caro",
        "cash_closing": 18799,
        "money_to_bank": 0,
        "paid_out": 0,
        "paid_in": 0,
        "cash_opening": 0,
        "cash_total": 18799,
        "gift_cards_total": 0,
        "mobilepay_total": 0,
        "cards_total": 92111.65,
        "payments_total": 110910.65,
        "sales_pos": 100292.54,
        "cash_difference": -10618.11,
        "cash_diff_flagged": True,
        "flagged_reason": "Cash difference -10,618.11 kr (short by 10,618.11, threshold 100)",
        "terminals": [
            {"terminal_name": "Front bar", "dankort": 24292.51, "teller": 31455.04, "amex": 0, "total": 55747.55},
            {"terminal_name": "Back bar",  "dankort": 17355.00, "teller": 19009.10, "amex": 0, "total": 36364.10},
            {"terminal_name": "Terrace",   "dankort": 0,        "teller": 0,        "amex": 0, "total": 0},
            {"terminal_name": "Takeaway",  "dankort": 0,        "teller": 0,        "amex": 0, "total": 0},
        ],
    }


def _load(xlsx_bytes: bytes):
    return load_workbook(io.BytesIO(xlsx_bytes), read_only=False, data_only=False)


def _all_label_values(ws) -> list[tuple[str, object]]:
    """Walk col A & B and return (label, value) pairs for inspection."""
    out = []
    for row in range(1, ws.max_row + 1):
        a = ws.cell(row=row, column=1).value
        b = ws.cell(row=row, column=2).value
        out.append((a, b))
    return out


# ─── Helper functions — pure, directly testable ────────────────────────

def test_money_value_basic():
    assert _money_value(14854) == 14854.0
    assert _money_value(14854.50) == 14854.5
    assert _money_value(0) == 0.0
    assert _money_value(-685.50) == -685.5


def test_money_value_handles_none():
    assert _money_value(None) is None


def test_money_value_handles_garbage():
    assert _money_value("not a number") is None
    assert _money_value([]) is None
    assert _money_value(object()) is None


def test_money_value_handles_string_number():
    """If upstream serialises a number as a string, coerce it."""
    assert _money_value("123.45") == 123.45
    assert _money_value("0") == 0.0


def test_danish_date_label_iso_input():
    date, day = _danish_date_label("2026-03-09")
    assert date == "09.03.2026"
    assert day == "Mandag"  # 9 March 2026 is a Monday


def test_danish_date_label_round_trip():
    date, day = _danish_date_label("9.3.2026 (Tirsdag)")
    assert date == "9.3.2026"
    assert day == "Tirsdag"


def test_danish_date_label_garbage_in():
    date, day = _danish_date_label("not-a-date")
    assert date == "not-a-date"
    assert day == ""


def test_danish_date_label_empty_uses_today():
    date, day = _danish_date_label("")
    # Should produce *some* DK-formatted date and day name
    assert "." in date
    assert day in {"Mandag", "Tirsdag", "Onsdag", "Torsdag", "Fredag", "Lørdag", "Søndag"}


# ─── Full xlsx render — basic contract ────────────────────────────────

def test_render_close_xlsx_returns_valid_xlsx_bytes():
    xlsx = render_close_xlsx(
        aggregated=_abigail_aggregated(),
        business_name="Mirabelle",
        date_label="9.3.2026 (Mandag)",
        currency="DKK",
    )
    assert isinstance(xlsx, bytes)
    # XLSX is a zip file → starts with PK
    assert xlsx.startswith(b"PK")
    assert len(xlsx) > 1000

    # Loads cleanly
    wb = _load(xlsx)
    assert wb is not None
    ws = wb.active
    assert ws is not None


def test_render_close_xlsx_handles_empty_aggregated():
    xlsx = render_close_xlsx(
        aggregated={},
        business_name="",
        date_label="",
        currency="DKK",
    )
    assert xlsx.startswith(b"PK")
    wb = _load(xlsx)
    # Should still have at least header rows
    assert wb.active.max_row > 0


def test_render_close_xlsx_handles_all_none_values():
    """Every numeric field None → render gracefully with blank cells
    (since _money_value returns None → openpyxl writes empty)."""
    xlsx = render_close_xlsx(
        aggregated={
            "cash_closing": None, "money_to_bank": None, "paid_out": None,
            "paid_in": None, "cash_opening": None, "cash_total": None,
            "gift_cards_total": None, "mobilepay_total": None,
            "cards_total": None, "payments_total": None,
            "sales_pos": None, "cash_difference": None,
            "cash_diff_flagged": False,
            "terminals": [],
        },
        business_name="Test",
        date_label="",
        currency="DKK",
    )
    assert xlsx.startswith(b"PK")
    wb = _load(xlsx)
    ws = wb.active
    # The "Cash closing - till out" row exists but value is empty
    rows = _all_label_values(ws)
    cash_close_row = next((r for r in rows if r[0] == "Cash closing - till out"), None)
    assert cash_close_row is not None
    assert cash_close_row[1] is None  # blank cell, not "—"


def test_render_close_xlsx_never_raises_on_malformed_input():
    """Defense layer — feed it deliberate garbage, expect an error xlsx,
    never a raise."""
    xlsx = render_close_xlsx(
        aggregated={"terminals": "not a list", "cash_closing": object()},  # type: ignore
        business_name="X",
        date_label="X",
        currency="DKK",
    )
    assert isinstance(xlsx, bytes)
    assert xlsx.startswith(b"PK")  # either main render or error fallback
    # Error fallback workbook should still be loadable
    wb = _load(xlsx)
    assert wb is not None


# ─── Full xlsx render — content assertions (the openpyxl advantage) ───

def test_render_close_xlsx_contains_mirabelle_labels():
    xlsx = render_close_xlsx(
        aggregated=_abigail_aggregated(),
        business_name="Mirabelle",
        date_label="9.3.2026 (Mandag)",
        currency="DKK",
    )
    wb = _load(xlsx)
    ws = wb.active
    labels = [c.value for c in ws["A"] if c.value]
    label_str = "  ".join(str(l) for l in labels)

    # Mirabelle row labels (the lingua franca — owner has used these for years)
    assert "Cash closing - till out" in label_str
    assert "Cash total" in label_str
    assert "Gift cards accepted (total)" in label_str
    assert "Mobile Pay" in label_str
    assert "Cards total" in label_str
    assert "Payments total" in label_str
    assert "Sales POS (incl. tax)" in label_str
    assert "Cash difference (+/-)" in label_str


def test_render_close_xlsx_business_name_in_sheet_title():
    xlsx = render_close_xlsx(
        aggregated=_abigail_aggregated(),
        business_name="Mirabelle",
        date_label="9.3.2026",
        currency="DKK",
    )
    wb = _load(xlsx)
    # Sheet title is the slugified business name
    assert "Mirabelle" in wb.sheetnames[0]


def test_render_close_xlsx_sheet_title_truncated_to_31_chars():
    """Excel sheet name limit is 31 chars — verify our slugifier respects it."""
    very_long_name = "A" * 60
    xlsx = render_close_xlsx(
        aggregated=_abigail_aggregated(),
        business_name=very_long_name,
        date_label="9.3.2026",
        currency="DKK",
    )
    wb = _load(xlsx)
    assert len(wb.sheetnames[0]) <= 31


def test_render_close_xlsx_sheet_title_strips_special_chars():
    """Slashes/colons in business names are illegal in Excel sheet names."""
    xlsx = render_close_xlsx(
        aggregated=_abigail_aggregated(),
        business_name="Café/Bar:K\\nightclub*",
        date_label="9.3.2026",
        currency="DKK",
    )
    wb = _load(xlsx)
    title = wb.sheetnames[0]
    for forbidden in ["/", "\\", ":", "*", "?"]:
        assert forbidden not in title


def test_render_close_xlsx_numeric_values_are_floats_not_strings():
    """The whole point of the Excel export: cells are real numbers,
    so customer can SUM them in their downstream workbook."""
    xlsx = render_close_xlsx(
        aggregated=_abigail_aggregated(),
        business_name="Mirabelle",
        date_label="9.3.2026",
        currency="DKK",
    )
    wb = _load(xlsx)
    ws = wb.active

    # Find the "Cash closing - till out" row, then check column B is a number
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "Cash closing - till out":
            val = ws.cell(row=row, column=2).value
            assert isinstance(val, (int, float))
            assert val == 18799.0
            break
    else:
        raise AssertionError("Cash closing - till out row not found")

    # Cards total should be a float, not "92.111,65 DKK"
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "Cards total":
            val = ws.cell(row=row, column=2).value
            assert isinstance(val, (int, float))
            assert abs(val - 92111.65) < 0.01
            break
    else:
        raise AssertionError("Cards total row not found")


def test_render_close_xlsx_cash_difference_negative_preserved():
    """Negative cash diff stays negative — owner needs to see the - sign
    AND the value should be a real negative number (not a string)."""
    xlsx = render_close_xlsx(
        aggregated=_abigail_aggregated(),
        business_name="Mirabelle",
        date_label="9.3.2026",
        currency="DKK",
    )
    wb = _load(xlsx)
    ws = wb.active
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "Cash difference (+/-)":
            val = ws.cell(row=row, column=2).value
            assert isinstance(val, (int, float))
            assert val < 0
            assert abs(val - (-10618.11)) < 0.01
            return
    raise AssertionError("Cash difference (+/-) row not found")


def test_render_close_xlsx_currency_in_number_format():
    """The number_format string carries the currency suffix so cells
    display '14.854,00 DKK' but underlying value stays numeric."""
    xlsx = render_close_xlsx(
        aggregated=_abigail_aggregated(),
        business_name="Mirabelle",
        date_label="9.3.2026",
        currency="DKK",
    )
    wb = _load(xlsx)
    ws = wb.active
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "Cash closing - till out":
            fmt = ws.cell(row=row, column=2).number_format
            assert "DKK" in fmt
            return
    raise AssertionError("Cash closing - till out row not found")


def test_render_close_xlsx_currency_override_passed_through():
    """Non-DKK customer (Vietnamese cafe, Norwegian) gets their currency
    in the number_format."""
    xlsx = render_close_xlsx(
        aggregated=_abigail_aggregated(),
        business_name="Vietnamese cafe",
        date_label="9.3.2026",
        currency="VND",
    )
    wb = _load(xlsx)
    ws = wb.active
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "Cash closing - till out":
            fmt = ws.cell(row=row, column=2).number_format
            assert "VND" in fmt
            return
    raise AssertionError("Cash closing - till out row not found")


def test_render_close_xlsx_per_terminal_blocks():
    """Each terminal gets its own Dankort/Teller/Amex/Total block, indexed
    by terminal_name (not just T1, T2, ...)."""
    xlsx = render_close_xlsx(
        aggregated=_abigail_aggregated(),
        business_name="Mirabelle",
        date_label="9.3.2026",
        currency="DKK",
    )
    wb = _load(xlsx)
    ws = wb.active
    labels = [str(c.value) for c in ws["A"] if c.value]
    label_str = "  ".join(labels)

    # Terminal name appears as a header row
    assert "Front bar" in label_str
    assert "Back bar" in label_str
    # Per-terminal sub-rows
    assert "Dankort from term. 1" in label_str
    assert "Teller from term. 1" in label_str
    assert "Amex from term. 1" in label_str
    assert "Total term. 1" in label_str
    # All four terminals
    assert "Total term. 4" in label_str


def test_render_close_xlsx_with_six_terminals():
    """6 terminals — biggest realistic restaurant — should render."""
    agg = _abigail_aggregated()
    agg["terminals"] = [
        {"terminal_name": f"Terminal {i}", "dankort": 1000, "teller": 2000, "amex": 0, "total": 3000}
        for i in range(1, 7)
    ]
    xlsx = render_close_xlsx(
        aggregated=agg,
        business_name="Big Restaurant",
        date_label="9.3.2026",
        currency="DKK",
    )
    wb = _load(xlsx)
    ws = wb.active
    labels = [str(c.value) for c in ws["A"] if c.value]
    label_str = "  ".join(labels)
    assert "Total term. 6" in label_str


def test_render_close_xlsx_single_terminal_business():
    """Single-terminal close (no terminals[] list) still renders cleanly."""
    agg = _abigail_aggregated()
    agg["terminals"] = []
    xlsx = render_close_xlsx(
        aggregated=agg,
        business_name="Tiny shop",
        date_label="9.3.2026",
        currency="DKK",
    )
    wb = _load(xlsx)
    # No terminal blocks but everything else is there
    ws = wb.active
    labels = [str(c.value) for c in ws["A"] if c.value]
    label_str = "  ".join(labels)
    assert "Cash closing - till out" in label_str
    assert "Cards total" in label_str


# ─── Compliance / footer / metadata fields ────────────────────────────

def test_render_close_xlsx_contains_bogforingsloven_footer():
    """Bogføringsloven §10 retention disclosure appears in footer."""
    xlsx = render_close_xlsx(
        aggregated=_abigail_aggregated(),
        business_name="Mirabelle",
        date_label="9.3.2026",
        currency="DKK",
    )
    wb = _load(xlsx)
    ws = wb.active
    found = False
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value
        if v and "Bogføringsloven §10" in str(v):
            found = True
            break
    assert found, "Bogføringsloven §10 footer missing"


def test_render_close_xlsx_with_business_profile():
    """CVR + DK address line passes through."""
    xlsx = render_close_xlsx(
        aggregated=_abigail_aggregated(),
        business_name="Restaurant Mirabelle ApS",
        date_label="9.3.2026 (Mandag)",
        currency="DKK",
        business_profile={
            "org_number": "44544891",
            "address": "Nørregade 12",
            "zipcode": "1165",
            "city": "København K",
            "country": "DK",
        },
    )
    wb = _load(xlsx)
    ws = wb.active
    found_cvr = False
    for row in range(1, 8):
        v = ws.cell(row=row, column=1).value
        if v and "CVR 44544891" in str(v):
            found_cvr = True
            break
    assert found_cvr, "CVR not in header"


def test_render_close_xlsx_with_bilagsnummer():
    xlsx = render_close_xlsx(
        aggregated=_abigail_aggregated(),
        business_name="Mirabelle",
        date_label="9.3.2026",
        currency="DKK",
        bilagsnummer="K2026-0042",
    )
    wb = _load(xlsx)
    ws = wb.active
    found = False
    for row in range(1, 10):
        v = ws.cell(row=row, column=1).value
        if v and "K2026-0042" in str(v):
            found = True
            break
    assert found, "Bilagsnummer not in header line"


def test_render_close_xlsx_no_business_profile_still_renders():
    xlsx = render_close_xlsx(
        aggregated=_abigail_aggregated(),
        business_name="Mirabelle",
        date_label="9.3.2026",
        currency="DKK",
        business_profile=None,
    )
    assert xlsx.startswith(b"PK")
    wb = _load(xlsx)
    assert wb.active.max_row > 0


# ─── MOMS section behaviour (mirrors PDF) ─────────────────────────────

def test_render_close_xlsx_moms_back_derived_from_payments_total():
    """Without revenue.* fields, MOMS back-derives from payments_total at 25%."""
    agg = _abigail_aggregated()
    agg.pop("revenue", None)
    xlsx = render_close_xlsx(
        aggregated=agg,
        business_name="Mirabelle",
        date_label="9.3.2026",
        currency="DKK",
    )
    wb = _load(xlsx)
    ws = wb.active
    # Section header should mention "estimat"
    found_estimat = False
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value
        if v and "MOMS" in str(v) and "estimat" in str(v):
            found_estimat = True
            break
    assert found_estimat, "MOMS estimat header missing on back-derive path"


def test_render_close_xlsx_moms_uses_supplied_revenue():
    """When revenue.* fields supplied, no estimat marker (real numbers)."""
    agg = _abigail_aggregated()
    agg["revenue"] = {
        "subtotal_excl_moms": 80234.12,
        "moms_amount": 20058.53,
        "total_incl_moms": 100292.65,
    }
    xlsx = render_close_xlsx(
        aggregated=agg,
        business_name="Mirabelle",
        date_label="9.3.2026",
        currency="DKK",
    )
    wb = _load(xlsx)
    ws = wb.active
    moms_header = None
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value
        if v and str(v).startswith("MOMS"):
            moms_header = str(v)
            break
    assert moms_header is not None
    assert "estimat" not in moms_header  # real numbers, not estimated


# ─── Cash-diff styling ────────────────────────────────────────────────

def test_render_close_xlsx_flagged_diff_uses_amber_fill():
    agg = _abigail_aggregated()
    agg["cash_diff_flagged"] = True
    xlsx = render_close_xlsx(
        aggregated=agg,
        business_name="Mirabelle",
        date_label="9.3.2026",
        currency="DKK",
    )
    wb = _load(xlsx)
    ws = wb.active
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "Cash difference (+/-)":
            fill = ws.cell(row=row, column=2).fill
            # Amber background hex (lowercase 'fef3c7' from the source)
            assert fill.start_color.rgb is not None
            assert "FEF3C7" in fill.start_color.rgb.upper()
            return
    raise AssertionError("Cash difference row not found")


def test_render_close_xlsx_clean_diff_uses_green_fill():
    agg = _abigail_aggregated()
    agg["cash_diff_flagged"] = False
    agg["flagged_reason"] = ""
    xlsx = render_close_xlsx(
        aggregated=agg,
        business_name="Mirabelle",
        date_label="9.3.2026",
        currency="DKK",
    )
    wb = _load(xlsx)
    ws = wb.active
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "Cash difference (+/-)":
            fill = ws.cell(row=row, column=2).fill
            # Green background hex (f0fdf4)
            assert fill.start_color.rgb is not None
            assert "F0FDF4" in fill.start_color.rgb.upper()
            return
    raise AssertionError("Cash difference row not found")


def test_render_close_xlsx_a4_portrait_page_setup():
    """Owner who prints the workbook gets A4 portrait, not Letter landscape."""
    xlsx = render_close_xlsx(
        aggregated=_abigail_aggregated(),
        business_name="Mirabelle",
        date_label="9.3.2026",
        currency="DKK",
    )
    wb = _load(xlsx)
    ws = wb.active
    assert ws.page_setup.orientation == ws.ORIENTATION_PORTRAIT
    # PAPERSIZE_A4 is stored as the string "9" on the class but round-trips
    # back from the XML as int 9 — so compare loosely as string.
    assert str(ws.page_setup.paperSize) == str(ws.PAPERSIZE_A4)
